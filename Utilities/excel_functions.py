import openpyxl
from openpyxl import load_workbook

class Excel_Operations:

    # Creating a constructor
    def __init__(self, file_name, sheet_number):
        self.filename = file_name
        self.sheet_number = sheet_number

    # Method to get max row
    def row_count(self):
        workbook = load_workbook(self.filename)
        sheet = workbook[self.sheet_number]
        return sheet.max_row

    # Method to get max column
    def column_count(self):
        workbook = load_workbook(self.filename)
        sheet = workbook[self.sheet_number]
        return sheet.max_column

    # Method to read data from excel
    def read_data(self,row_number,column_number):
        workbook = load_workbook(self.filename)
        sheet = workbook[self.sheet_number]
        return sheet.cell(row=row_number,column=column_number).value

    # # Method to write data in excel
    def write_data(self,row_number,column_number,data):
        workbook = load_workbook(self.filename)
        sheet = workbook[self.sheet_number]
        sheet.cell(row=row_number,column=column_number).value = data
        workbook.save(self.filename)

# if __name__=="__main__":
#     excel_file = "E:\\Workspace\\GUVI\\PAT-Tasks\\PAT-Task-12\\Data\\test_data.xlsx"
#     sheet_number = "Sheet1"
#     data = [1,2,3,4,5]
#
#     excel_obj = Excel_Operations(excel_file,sheet_number)
#     row = excel_obj.row_count()
#     print(excel_obj.row_count())
#     print(excel_obj.column_count())
#     print(excel_obj.read_data(1,4))
#     # for row in range(2, row+1):
#     #     for i in range(len(data)):
#     #         excel_obj.write_data(row_number=row,column_number=8,data=data[i])
#     start_row = 2
#     for i in range(len(data)):
#         excel_obj.write_data(row_number=start_row + i, column_number=8, data=data[i])



